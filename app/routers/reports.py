from fastapi import APIRouter, Depends
from sqlalchemy.orm import Session
from sqlalchemy import func, extract
from datetime import date
from fastapi.responses import FileResponse
from openpyxl import Workbook
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet
import tempfile

from app.database import get_db
from app.models import Collection, Chair, Store, Employee, Alert
from app.routers import settings

router = APIRouter(prefix="/api/reports", tags=["Reports"])


@router.get("/monthly")
def monthly_report(
    month: int,
    year: int,
    db: Session = Depends(get_db)
):
    total_revenue = (
        db.query(func.sum(Collection.total_amount))
        .filter(
            extract("month", Collection.date) == month,
            extract("year", Collection.date) == year,
        )
        .scalar()
        or 0
    )

    # Temporary revenue split
    company_share = 0
    store_share = 0

    stores = db.query(Store).all()

    for store in stores:

        revenue = (
            db.query(func.sum(Collection.total_amount))
            .filter(
                Collection.store_id == store.id,
                extract("month", Collection.date) == month,
                extract("year", Collection.date) == year,
            )
            .scalar()
            or 0
        )

        company_share += revenue * 0.75
        store_share += revenue * 0.25

    # Best Chair
    best_chair = (
        db.query(
            Chair.device_id,
            func.sum(Collection.total_amount).label("revenue")
        )
        .join(Collection, Collection.chair_id == Chair.id)
        .filter(
            extract("month", Collection.date) == month,
            extract("year", Collection.date) == year,
        )
        .group_by(Chair.device_id)
        .order_by(func.sum(Collection.total_amount).desc())
        .first()
    )

    # Lowest Chair
    lowest_chair = (
        db.query(
            Chair.device_id,
            func.sum(Collection.total_amount).label("revenue")
        )
        .join(Collection, Collection.chair_id == Chair.id)
        .filter(
            extract("month", Collection.date) == month,
            extract("year", Collection.date) == year,
        )
        .group_by(Chair.device_id)
        .order_by(func.sum(Collection.total_amount).asc())
        .first()
    )

    # Best Store
    best_store = (
        db.query(
            Store.name,
            func.sum(Collection.total_amount).label("revenue")
        )
        .join(Collection, Collection.store_id == Store.id)
        .filter(
            extract("month", Collection.date) == month,
            extract("year", Collection.date) == year,
        )
        .group_by(Store.name)
        .order_by(func.sum(Collection.total_amount).desc())
        .first()
    )

    # Revenue Trend
    revenue_trend = []

    for m in range(max(1, month - 5), month + 1):

        revenue = (
            db.query(func.sum(Collection.total_amount))
            .filter(
                extract("month", Collection.date) == m,
                extract("year", Collection.date) == year,
            )
            .scalar()
            or 0
        )

        revenue_trend.append(
            {
                "month": m,
                "revenue": round(revenue, 2)
            }
        )

    # Sales Team Performance
    sales_team = []

    employees = db.query(Employee).all()

    for employee in employees:

        revenue = (
            db.query(func.sum(Collection.total_amount))
            .join(Chair, Chair.id == Collection.chair_id)
            .filter(
                Chair.installed_by_employee_id == employee.id
            )
            .scalar()
            or 0
        )

        sales_team.append(
            {
                "employee": employee.name,
                "revenue": round(revenue, 2)
            }
        )

    # Store Breakdown
    store_breakdown = []

    stores = db.query(Store).all()

    for store in stores:

        revenue = (
            db.query(func.sum(Collection.total_amount))
            .filter(
                Collection.store_id == store.id,
                extract("month", Collection.date) == month,
                extract("year", Collection.date) == year,
            )
            .scalar()
            or 0
        )

        chair_count = (
            db.query(Chair)
            .filter(Chair.store_id == store.id)
            .count()
        )

        store_breakdown.append(
            {
                "store_name": store.name,
                "city": store.city,
                "chairs": chair_count,
                "gross": round(revenue, 2),
                "company": round(revenue * 0.75, 2),
                "partner": round(revenue * 0.25, 2),
            }
        )

    alerts_summary = {
        "critical":
            db.query(Alert)
            .filter(Alert.severity == "critical")
            .count(),

        "warning":
            db.query(Alert)
            .filter(Alert.severity == "warning")
            .count(),
    }

    return {
        "summary": {
            "month": f"{month}/{year}",
            "total_revenue": round(total_revenue, 2),
            "company_share": round(company_share, 2),
            "store_share": round(store_share, 2),
        },

        "best_store": {
            "name": best_store.name if best_store else None,
            "revenue": float(best_store.revenue) if best_store else 0,
        },

        "best_chair": {
            "device_id": best_chair.device_id if best_chair else None,
            "revenue": float(best_chair.revenue) if best_chair else 0,
        },

        "lowest_chair": {
            "device_id": lowest_chair.device_id if lowest_chair else None,
            "revenue": float(lowest_chair.revenue) if lowest_chair else 0,
        },

        "revenue_trend": revenue_trend,

        "sales_team_performance": sales_team,

        "store_breakdown": store_breakdown,

        "alerts_summary": alerts_summary,
    }

@router.get("/export/excel")
def export_excel(
    db: Session = Depends(get_db)
):
    wb = Workbook()

    # Sheet 1
    ws = wb.active
    ws.title = "Executive Summary"

    total_revenue = (
        db.query(
            func.sum(
                Collection.total_amount
            )
        )
        .scalar()
        or 0
    )

    company_share = (total_revenue * settings.company_share_percentage)/100



    store_share = (total_revenue * settings.store_share_percentage)/100

    ws.append(["Metric", "Value"])
    ws.append(["Total Revenue", float(total_revenue)])
    ws.append(["Company Share", float(company_share)])
    ws.append(["Store Share", float(store_share)])

    # Sheet 2 - Stores
    store_sheet = wb.create_sheet(
        "Stores"
    )

    store_sheet.append(
        [
            "Store",
            "City",
            "Revenue"
        ]
    )

    stores = db.query(
        Store
    ).all()

    for store in stores:

        revenue = (
            db.query(
                func.sum(
                    Collection.total_amount
                )
            )
            .filter(
                Collection.store_id
                == store.id
            )
            .scalar()
            or 0
        )

        store_sheet.append(
            [
                store.name,
                store.city,
                float(revenue)
            ]
        )

    # Sheet 3 - Employees
    employee_sheet = wb.create_sheet(
        "Employees"
    )

    employee_sheet.append(
        [
            "Employee Name",
            "Designation"
        ]
    )

    employees = db.query(
        Employee
    ).all()

    for emp in employees:

        employee_sheet.append(
            [
                emp.name,
                emp.designation
            ]
        )

    # Sheet 4 - Alerts
    alert_sheet = wb.create_sheet(
        "Alerts"
    )

    alert_sheet.append(
        [
            "Critical",
            "Warning"
        ]
    )

    critical_count = (
        db.query(Alert)
        .filter(
            Alert.severity == "critical"
        )
        .count()
    )

    warning_count = (
        db.query(Alert)
        .filter(
            Alert.severity == "warning"
        )
        .count()
    )

    alert_sheet.append(
        [
            critical_count,
            warning_count
        ]
    )

    temp_file = tempfile.NamedTemporaryFile(
        delete=False,
        suffix=".xlsx"
    )

    wb.save(
        temp_file.name
    )

    return FileResponse(
        temp_file.name,
        media_type=(
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        ),
        filename="monthly_report.xlsx"
    )

@router.get("/export/pdf")
def export_pdf(
    db: Session = Depends(get_db)
):
    temp_file = tempfile.NamedTemporaryFile(
        delete=False,
        suffix=".pdf"
    )

    doc = SimpleDocTemplate(
        temp_file.name
    )

    styles = (
        getSampleStyleSheet()
    )

    content = []

    # Title
    content.append(
        Paragraph(
            "MasaZ Monthly Report",
            styles["Title"]
        )
    )

    content.append(
        Spacer(
            1,
            20
        )
    )

    # Revenue Summary
    total_revenue = (
        db.query(
            func.sum(
                Collection.total_amount
            )
        )
        .scalar()
        or 0
    )

    company_share = (total_revenue * settings.company_share_percentage)/100



    store_share = (total_revenue * settings.store_share_percentage)/100

    content.append(
        Paragraph(
            f"Total Revenue: ₹{round(total_revenue, 2)}",
            styles["Normal"]
        )
    )

    content.append(
        Paragraph(
            f"Company Share: ₹{round(company_share, 2)}",
            styles["Normal"]
        )
    )

    content.append(
        Paragraph(
            f"Store Share: ₹{round(store_share, 2)}",
            styles["Normal"]
        )
    )

    content.append(
        Spacer(
            1,
            15
        )
    )

    # Best Store
    best_store = (
        db.query(
            Store.name,
            func.sum(
                Collection.total_amount
            ).label(
                "revenue"
            )
        )
        .join(
            Collection,
            Collection.store_id
            == Store.id
        )
        .group_by(
            Store.name
        )
        .order_by(
            func.sum(
                Collection.total_amount
            ).desc()
        )
        .first()
    )

    if best_store:

        content.append(
            Paragraph(
                f"Best Store: {best_store.name}",
                styles["Heading2"]
            )
        )

        content.append(
            Paragraph(
                f"Revenue: ₹{round(best_store.revenue, 2)}",
                styles["Normal"]
            )
        )

    content.append(
        Spacer(
            1,
            15
        )
    )

    # Employees
    content.append(
        Paragraph(
            "Employees",
            styles["Heading2"]
        )
    )

    employees = db.query(
        Employee
    ).all()

    for emp in employees:

        content.append(
            Paragraph(
                f"{emp.name} - {emp.designation}",
                styles["Normal"]
            )
        )

    content.append(
        Spacer(
            1,
            15
        )
    )

    # Alerts
    critical_count = (
        db.query(Alert)
        .filter(
            Alert.severity == "critical"
        )
        .count()
    )

    warning_count = (
        db.query(Alert)
        .filter(
            Alert.severity == "warning"
        )
        .count()
    )

    content.append(
        Paragraph(
            "Alerts Summary",
            styles["Heading2"]
        )
    )

    content.append(
        Paragraph(
            f"Critical Alerts: {critical_count}",
            styles["Normal"]
        )
    )

    content.append(
        Paragraph(
            f"Warning Alerts: {warning_count}",
            styles["Normal"]
        )
    )

    doc.build(
        content
    )

    return FileResponse(
        temp_file.name,
        media_type="application/pdf",
        filename="monthly_report.pdf"
    )


